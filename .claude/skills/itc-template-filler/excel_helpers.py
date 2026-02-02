#!/usr/bin/env python3
"""
Excel Template Helper Functions
Provides utilities for safely working with Excel templates including merged cells
"""

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from pathlib import Path


def set_cell_value(sheet, cell_ref, value):
    """
    Safely set cell value, handling both regular and merged cells.

    Args:
        sheet: openpyxl worksheet object
        cell_ref: Cell reference string (e.g., 'E11', 'D4')
        value: Value to set (str, int, float, datetime, etc.)
    """
    cell = sheet[cell_ref]

    # Check if cell is part of a merged range
    if isinstance(cell, MergedCell):
        # Find the merged range containing this cell
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Get the top-left cell (the "real" cell)
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_cell = sheet.cell(min_row, min_col)
                top_left_cell.value = value
                return
        # If we get here, something is wrong
        raise ValueError(f"Cell {cell_ref} is a MergedCell but not in any merged range")
    else:
        # Regular cell - direct assignment works
        cell.value = value


def get_merged_cell_value(sheet, cell_ref):
    """
    Safely read value from merged or regular cell.

    Args:
        sheet: openpyxl worksheet object
        cell_ref: Cell reference string

    Returns:
        Cell value
    """
    cell = sheet[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                min_col, min_row, max_col, max_row = merged_range.bounds
                return sheet.cell(min_row, min_col).value
    return cell.value


def inspect_merged_cells(sheet):
    """
    Print all merged cell ranges in sheet.

    Args:
        sheet: openpyxl worksheet object
    """
    print(f"Merged cells in '{sheet.title}':")
    for merged_range in sheet.merged_cells.ranges:
        print(f"  {merged_range}")


def inspect_template_structure(workbook):
    """
    Inspect template to understand structure.

    Args:
        workbook: openpyxl workbook object

    Returns:
        dict: Template structure information
    """
    info = {
        'sheets': [],
        'merged_cells': {},
        'formulas': {},
        'data_validation': {}
    }

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        info['sheets'].append(sheet_name)

        # Find merged cells
        merged = [str(mr) for mr in sheet.merged_cells.ranges]
        if merged:
            info['merged_cells'][sheet_name] = merged

        # Find formula cells (don't overwrite these!)
        formulas = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and cell.data_type == 'f':
                    formulas.append(f"{cell.coordinate}: {cell.value}")
        if formulas:
            info['formulas'][sheet_name] = formulas

    return info


def print_inspection_report(report):
    """
    Pretty print inspection report.

    Args:
        report: dict from inspect_template_structure()
    """
    print("=" * 80)
    print(f"TEMPLATE INSPECTION")
    print("=" * 80)

    for sheet_name in report['sheets']:
        print(f"\n📄 Sheet: {sheet_name}")

        if sheet_name in report.get('merged_cells', {}):
            merged = report['merged_cells'][sheet_name]
            print(f"\n   🔗 Merged Cells ({len(merged)})")
            for mc in merged[:5]:  # Show first 5
                print(f"      {mc}")
            if len(merged) > 5:
                print(f"      ... and {len(merged) - 5} more")

        if sheet_name in report.get('formulas', {}):
            formulas = report['formulas'][sheet_name]
            print(f"\n   📊 Formula Cells ({len(formulas)})")
            for fc in formulas[:3]:
                print(f"      {fc}")
            if len(formulas) > 3:
                print(f"      ... and {len(formulas) - 3} more")

    print("\n" + "=" * 80)


def validate_template(template_path):
    """
    Validate that template exists and can be loaded.

    Args:
        template_path: Path to Excel template

    Returns:
        tuple: (success: bool, workbook or None, error_msg or None)
    """
    template_path = Path(template_path)

    if not template_path.exists():
        return False, None, f"Template not found: {template_path}"

    if not template_path.suffix in ['.xlsx', '.xlsm']:
        return False, None, f"Invalid file type: {template_path.suffix} (expected .xlsx or .xlsm)"

    try:
        # Use keep_vba=True for .xlsm files to preserve macros
        if template_path.suffix == '.xlsm':
            workbook = openpyxl.load_workbook(template_path, keep_vba=True)
        else:
            workbook = openpyxl.load_workbook(template_path)
        return True, workbook, None
    except Exception as e:
        return False, None, f"Error loading template: {e}"


def create_backup(file_path):
    """
    Create backup copy of file before modification.

    Args:
        file_path: Path to file

    Returns:
        Path: Backup file path
    """
    file_path = Path(file_path)
    backup_path = file_path.parent / f"{file_path.stem}_backup{file_path.suffix}"

    import shutil
    shutil.copy2(file_path, backup_path)

    return backup_path
